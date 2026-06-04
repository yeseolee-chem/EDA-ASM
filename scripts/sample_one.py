"""Sample 1 random reaction from the stage5a 500-reaction set.

Records (seed, chosen reaction_id) under data/selection/ so re-runs are
deterministic and the full 500-run replays the same first-pick.

Usage:
    python scripts/sample_one.py [--seed 20260513]
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl


REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "reports" / "stage5a_500_sample.xlsx"
STAGE5A_DIR = REPO / "outputs" / "stage5a" / "per_reaction"
SEL_DIR = REPO / "data" / "selection"


def load_reactions() -> list[dict]:
    wb = openpyxl.load_workbook(str(XLSX), read_only=True)
    ws = wb["reactions"]
    header = [c.value for c in next(ws.iter_rows())]
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, r)))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=20260513,
                    help="numpy RNG seed (default: 20260513)")
    args = ap.parse_args()

    SEL_DIR.mkdir(parents=True, exist_ok=True)

    rxns = load_reactions()
    print(f"Total reactions in xlsx: {len(rxns)}")

    # Keep only reactions whose stage5a per_reaction folder + result.json exist
    present = []
    for r in rxns:
        rid = r["reaction_id"]
        if (STAGE5A_DIR / rid / "result.json").exists():
            present.append(r)
    print(f"With result.json on disk: {len(present)}")

    rng = np.random.default_rng(args.seed)
    idx = int(rng.integers(0, len(present)))
    chosen = present[idx]

    # Write rxn_id_list.txt (full 500) for later SLURM array
    list_path = REPO / "outputs" / "stage5b" / "rxn_id_list.txt"
    list_path.parent.mkdir(parents=True, exist_ok=True)
    with open(list_path, "w") as f:
        for r in present:
            f.write(r["reaction_id"] + "\n")
    print(f"Wrote {list_path} ({len(present)} ids)")

    # Persist the choice
    record = {
        "seed": args.seed,
        "n_total": len(rxns),
        "n_available": len(present),
        "chosen_index_into_available": idx,
        "chosen": {
            "reaction_id": chosen["reaction_id"],
            "pattern": chosen.get("pattern"),
            "p2_subtype": chosen.get("p2_subtype"),
            "formula": chosen.get("formula"),
            "n_atoms": chosen.get("n_atoms"),
            "n_heavy_atoms": chosen.get("n_heavy_atoms"),
            "n_fragments": chosen.get("n_fragments"),
            "activation_energy_eV": chosen.get("activation_energy"),
            "energy_R": chosen.get("energy_R"),
            "energy_TS": chosen.get("energy_TS"),
            "energy_P": chosen.get("energy_P"),
            "review_status": chosen.get("review_status"),
        },
        "selection_timestamp_utc": datetime.now(timezone.utc).isoformat(),
    }
    sel_path = SEL_DIR / "first_pick.json"
    with open(sel_path, "w") as f:
        json.dump(record, f, indent=2)
    print(f"\nChosen reaction:")
    for k, v in record["chosen"].items():
        print(f"  {k:24s} {v}")
    print(f"\nSaved to {sel_path}")


if __name__ == "__main__":
    main()
